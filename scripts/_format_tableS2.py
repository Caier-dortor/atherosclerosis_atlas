"""Format Table S2 as submission-ready Excel with proper headers, grouping, and annotations."""
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

TABLE_DIR = Path("D:/openclaw_workspace/atherosclerosis_atlas/results/supplementary")
df = pd.read_csv(TABLE_DIR / "TableS2_effect_sizes_bootstrap.csv")
boot_ci = pd.read_csv(TABLE_DIR / "TableS2_femoral_bootstrap_ci.csv")

# ── Module grouping (for visual separation) ──
MODULE_GROUPS = {
    "TI Composite": ["TI_composite"],
    "TI Sub-scores": ["TI_Inflammation_score", "TI_Metabolic_score", "TI_PRR_score",
                       "TI_H3K27ac_score", "TI_H3K4me3_score", "TI_HDAC_SIRT_score"],
    "Macrophage Subtype Scores": ["Foamy_Mac_score", "Inflammatory_Mac_score",
                                   "HMOX1_Mac_score", "Resident_Mac_score"],
    "Metabolic Scores": ["FAO_score", "FAS_score", "Glycolysis_score", "OXPHOS_score",
                          "Cholesterol_score", "Hypoxia_score"],
    "Inflammation Scores": ["Acute_Inflammation_score"],
    "PVAT Scores": ["Healthy_PVAT_score", "Disease_PVAT_score"],
}
GROUP_ORDER = ["TI Composite", "TI Sub-scores", "Macrophage Subtype Scores",
               "Metabolic Scores", "Inflammation Scores", "PVAT Scores"]

# ── Create workbook ──
wb = Workbook()
ws = wb.active
ws.title = "Table S2"

# Styles
header_font = Font(name="Times New Roman", size=8, bold=True)
group_font = Font(name="Times New Roman", size=8, bold=True, color="1F4E79")
data_font = Font(name="Times New Roman", size=7.5)
note_font = Font(name="Times New Roman", size=7, italic=True, color="666666")
sig_font = Font(name="Times New Roman", size=7.5, bold=True, color="CC0000")
thin_border = Border(
    bottom=Side(style="hair", color="CCCCCC"),
)
group_border = Border(
    bottom=Side(style="thin", color="1F4E79"),
)
header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
wrap_align = Alignment(wrap_text=True, vertical="center", horizontal="center")
left_align = Alignment(wrap_text=True, vertical="center", horizontal="left")

# ── Column layout ──
COLUMNS = [
    ("Module", 28),
    ("Comparison", 22),
    ("H", 8),
    ("KW p", 10),
    ("eta^2", 8),
    ("MW U", 10),
    ("MW p", 10),
    ("Carotid\nMean", 10),
    ("Coronary\nMean", 10),
    ("N\n(Car)", 6),
    ("N\n(Cor)", 6),
    ("Cohen's d", 10),
    ("Hedges' g", 10),
    ("Femoral\nMean", 10),
    ("N\n(Fem)", 6),
    ("MW p\n(Bonf.)", 10),
    ("Sig.", 5),
]

# Write header row
for j, (col_name, width) in enumerate(COLUMNS, 1):
    cell = ws.cell(row=1, column=j, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap_align
    cell.border = Border(bottom=Side(style="thin", color="999999"))
    ws.column_dimensions[get_column_letter(j)].width = width

ws.row_dimensions[1].height = 28

# ── Data rows, grouped ──
row = 2
COMP_ORDER = ["carotid vs coronary", "carotid vs femoral", "coronary vs femoral"]

# Build lookup
df_lookup = {}
for _, r in df.iterrows():
    df_lookup[(r["Module"], r["Comparison"])] = r

def fmt_p(v, low=1e-4):
    if pd.isna(v): return ""
    if v < 1e-10: return "p<1e-10"
    if v < low: return f"{v:.2e}"
    return f"{v:.4f}"

def fmt_d(v, decimals=3):
    if pd.isna(v): return ""
    return f"{v:.{decimals}f}"

for grp_name in GROUP_ORDER:
    modules = MODULE_GROUPS[grp_name]
    # Group header
    cell = ws.cell(row=row, column=1, value=grp_name.upper())
    cell.font = group_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
    for j in range(1, len(COLUMNS) + 1):
        ws.cell(row=row, column=j).border = group_border
    ws.row_dimensions[row].height = 18
    row += 1

    for module in modules:
        for comp in COMP_ORDER:
            if (module, comp) not in df_lookup:
                continue
            r = df_lookup[(module, comp)]
            vals = [
                module.replace("_score", "").replace("_", " "),
                comp,
                fmt_d(r["H_statistic"], 2),
                fmt_p(r["KW_p"]),
                fmt_d(r["eta_squared"], 4),
                fmt_d(r["MW_U"], 1),
                fmt_p(r["MW_p"]),
                fmt_d(r["carotid_mean"], 4),
                fmt_d(r["coronary_mean"], 4),
                str(int(r["carotid_n"])) if not pd.isna(r["carotid_n"]) else "",
                str(int(r["coronary_n"])) if not pd.isna(r["coronary_n"]) else "",
                fmt_d(r["Cohens_d"], 3),
                fmt_d(r["Hedges_g"], 3),
                fmt_d(r["femoral_mean"], 4),
                str(int(r["femoral_n"])) if not pd.isna(r["femoral_n"]) else "",
                fmt_p(r["MW_p_bonf"]),
                str(r["Significance"]),
            ]
            for j, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=j, value=val)
                cell.font = data_font
                cell.alignment = Alignment(vertical="center", horizontal="center" if j > 1 else "left")
                cell.border = thin_border
                # Bold significant rows
                if r["Significance"] in ("***", "**", "*"):
                    if j in (4, 7, 16, 17):  # p-value + significance columns
                        cell.font = sig_font
            ws.row_dimensions[row].height = 15
            row += 1

    # Blank separator row
    row += 1

# ── Add footnote rows ──
row += 1
footnotes = [
    "Table S2. Effect sizes and bootstrap confidence intervals for module score comparisons across vascular beds.",
    "Statistical tests: Kruskal-Wallis (KW) omnibus test across 3 beds; Mann-Whitney U for pairwise comparisons.",
    "Effect sizes: eta-squared (eta^2) for KW; Cohen's d and Hedges' g for pairwise MW comparisons.",
    "MW p (Bonf.): Bonferroni-corrected Mann-Whitney p-values (3 comparisons per module).",
    "Significance: *** p<0.001, ** p<0.01, * p<0.05, ns = not significant (after Bonferroni correction).",
    "Femoral artery n=7 donors; carotid n=50; coronary n=13. See Table S2b for femoral bootstrap CIs.",
    "",
    f"Module scores were computed as mean expression of gene sets from MSigDB Hallmark/GO terms and custom trained-immunity signatures.",
]
for fn in footnotes:
    cell = ws.cell(row=row, column=1, value=fn)
    cell.font = note_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
    ws.row_dimensions[row].height = 14
    row += 1

# ── Sheet 2: Femoral Bootstrap CI ──
ws2 = wb.create_sheet("Table S2b Femoral Bootstrap CI")
ws2_headers = ["Module", "Femoral Mean", "95% CI Lower", "95% CI Upper", "Carotid Mean", "Femoral N", "Carotid N"]
for j, h in enumerate(ws2_headers, 1):
    cell = ws2.cell(row=1, column=j, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap_align
    cell.border = Border(bottom=Side(style="thin", color="999999"))
    ws2.column_dimensions[get_column_letter(j)].width = max(len(h) + 4, 18)

ws2.row_dimensions[1].height = 22

for i, (_, r) in enumerate(boot_ci.iterrows()):
    row_i = i + 2
    vals = [
        r["Module"].replace("_score", "").replace("_", " "),
        f'{r["Femoral_mean"]:.4f}',
        f'{r["Femoral_CI95_lo"]:.4f}',
        f'{r["Femoral_CI95_hi"]:.4f}',
        f'{r["Carotid_mean"]:.4f}',
        str(int(r["Femoral_n"])),
        str(int(r["Carotid_n"])),
    ]
    for j, val in enumerate(vals, 1):
        cell = ws2.cell(row=row_i, column=j, value=val)
        cell.font = data_font
        cell.alignment = Alignment(vertical="center", horizontal="center" if j > 1 else "left")
        cell.border = thin_border
    ws2.row_dimensions[row_i].height = 15

# Add footnote
row2 = len(boot_ci) + 3
fn2 = "Bootstrap CIs computed with 10,000 iterations, resampling femoral donors with replacement."
cell = ws2.cell(row=row2, column=1, value=fn2)
cell.font = note_font
ws2.merge_cells(start_row=row2, start_column=1, end_row=row2, end_column=len(ws2_headers))

# ── Freeze panes ──
ws.freeze_panes = "A2"
ws2.freeze_panes = "A2"

# ── Print settings ──
ws.sheet_properties.pageSetUpPr = None
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws2.page_setup.orientation = "landscape"

# ── Save ──
out = TABLE_DIR / "TableS2_effect_sizes_formatted.xlsx"
wb.save(out)
print(f"Saved: {out}")