"""Table S2: Add bootstrap CI for effect sizes + power analysis. Then re-format final Excel."""
import pandas as pd
import numpy as np
from pathlib import Path
from scipy.stats import mannwhitneyu, norm
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

TABLE_DIR = Path("D:/openclaw_workspace/atherosclerosis_atlas/results/supplementary")
np.random.seed(42)

# ── Load donor data ──
donor = pd.read_csv("D:/openclaw_workspace/atherosclerosis_atlas/results/donor_level_scores.csv")
df_old = pd.read_csv(TABLE_DIR / "TableS2_effect_sizes_bootstrap.csv")

# Module list
MODULES = df_old['Module'].unique().tolist()
BEDS = {'carotid': 50, 'coronary': 13, 'femoral': 7}

# ── Helper: Cohen's d from raw data ──
def cohens_d(x, y):
    nx, ny = len(x), len(y)
    if nx < 2 or ny < 2:
        return np.nan
    # Pooled SD
    sp = np.sqrt(((nx - 1) * np.var(x, ddof=1) + (ny - 1) * np.var(y, ddof=1)) / (nx + ny - 2))
    if sp < 1e-12:
        return 0.0
    return (np.mean(x) - np.mean(y)) / sp

def hedges_g(d, nx, ny):
    # Correction factor J = 1 - 3/(4*df - 1)
    df = nx + ny - 2
    if df < 1:
        return d
    J = 1 - 3 / (4 * df - 1)
    return d * J

# ── Bootstrap CI for Cohen's d ──
def bootstrap_cohens_d(x, y, n_boot=2000):
    nx, ny = len(x), len(y)
    d_obs = cohens_d(x, y)
    boot_vals = []
    for _ in range(n_boot):
        xb = np.random.choice(x, size=nx, replace=True)
        yb = np.random.choice(y, size=ny, replace=True)
        boot_vals.append(cohens_d(xb, yb))
    boot_vals = np.array(boot_vals)
    boot_vals = boot_vals[~np.isnan(boot_vals)]
    if len(boot_vals) == 0:
        return d_obs, np.nan, np.nan
    ci_lo = np.percentile(boot_vals, 2.5)
    ci_hi = np.percentile(boot_vals, 97.5)
    return d_obs, ci_lo, ci_hi

# ── Power analysis ──
def achieved_power(d, n1, n2, alpha=0.05):
    """Post-hoc power for two-sample t-test with given Cohen's d."""
    if pd.isna(d) or n1 < 2 or n2 < 2:
        return np.nan
    # Non-centrality parameter
    ncp = abs(d) * np.sqrt(n1 * n2 / (n1 + n2))
    df = n1 + n2 - 2
    # Critical t
    t_crit = norm.ppf(1 - alpha / 2)
    # Power = P(|t| > t_crit) under non-central t → approximate with normal
    power = norm.cdf(-t_crit, loc=ncp, scale=1) + (1 - norm.cdf(t_crit, loc=ncp, scale=1))
    return power

# ═══════════════════════════════════════════
# Compute bootstrap CI + power for each comparison
# ═══════════════════════════════════════════
print("Computing bootstrap CIs and power for all comparisons...")
CI_AND_POWER = {}

for module in MODULES:
    for comp in df_old[df_old['Module'] == module]['Comparison'].unique():
        parts = comp.split(' vs ')
        bed1, bed2 = parts[0], parts[1]

        vals1 = donor.loc[donor['plaque_location'] == bed1, module].dropna().values
        vals2 = donor.loc[donor['plaque_location'] == bed2, module].dropna().values
        n1, n2 = len(vals1), len(vals2)

        if n1 < 3 or n2 < 3:
            CI_AND_POWER[(module, comp)] = {
                'd_boot': np.nan, 'd_ci_lo': np.nan, 'd_ci_hi': np.nan,
                'power_0_5': np.nan, 'power_0_8': np.nan, 'min_detectable_d': np.nan
            }
            continue

        # Bootstrap CI for Cohen's d
        d_obs, d_ci_lo, d_ci_hi = bootstrap_cohens_d(vals1, vals2, n_boot=2000)

        # Achieved power
        power = achieved_power(d_obs, n1, n2)

        # Minimum detectable d at 80% power
        # d_min = (Z_alpha/2 + Z_beta) / sqrt(n1*n2/(n1+n2))
        z_alpha = norm.ppf(0.975)  # 1.96
        z_beta = norm.ppf(0.80)     # 0.84
        min_d = (z_alpha + z_beta) / np.sqrt(n1 * n2 / (n1 + n2))

        CI_AND_POWER[(module, comp)] = {
            'd_boot': d_obs, 'd_ci_lo': d_ci_lo, 'd_ci_hi': d_ci_hi,
            'power': power, 'min_detectable_d': min_d
        }

# ═══════════════════════════════════════════
# Build formatted Excel with all data
# ═══════════════════════════════════════════
print("Building formatted Excel...")

header_font = Font(name="Times New Roman", size=7.5, bold=True)
data_font = Font(name="Times New Roman", size=7)
note_font = Font(name="Times New Roman", size=6.5, italic=True, color="666666")
group_font = Font(name="Times New Roman", size=8, bold=True, color="1F4E79")
header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
thin_border = Border(bottom=Side(style="hair", color="CCCCCC"))
group_border = Border(bottom=Side(style="thin", color="1F4E79"))
center_align = Alignment(vertical="center", horizontal="center", wrap_text=True)
left_align = Alignment(vertical="center", horizontal="left", wrap_text=True)

MODULE_GROUPS = {
    "TI Composite": ["TI_composite"],
    "TI Sub-scores": ["TI_Inflammation_score", "TI_Metabolic_score", "TI_PRR_score",
                       "TI_H3K27ac_score", "TI_H3K4me3_score", "TI_HDAC_SIRT_score"],
    "Macrophage Subtype Scores": ["Foamy_Mac_score", "Inflammatory_Mac_score",
                                   "HMOX1_Mac_score", "Resident_Mac_score"],
    "Metabolic Scores": ["FAO_score", "FAS_score", "Glycolysis_score", "OXPHOS_score",
                          "Cholesterol_score", "Hypoxia_score"],
    "Inflammation Scores": ["Acute_Inflammation_score"],
    "PVAT Scores": ["Healthy_PVAT_score", "Disease_PVAT_score"],
}
GROUP_ORDER = list(MODULE_GROUPS.keys())
COMP_ORDER = ["carotid vs coronary", "carotid vs femoral", "coronary vs femoral"]

# Lookup
df_lookup = {}
for _, r in df_old.iterrows():
    df_lookup[(r["Module"], r["Comparison"])] = r

COL_HEADERS = [
    "Module", "Comparison",
    "H", "KW p", "eta^2",
    "MW U", "MW p", "MW p\n(Bonf.)",
    "Cohen's d", "d 95%CI lo", "d 95%CI hi",
    "Hedges' g",
    "Carotid\nMean", "Coronary\nMean", "Femoral\nMean",
    "N(Car)", "N(Cor)", "N(Fem)",
    "Achieved\nPower", "Min. Detect.\nd (80% power)", "Sig.",
]

wb = Workbook()
ws = wb.active
ws.title = "Table S2"

# Header row
for j, h in enumerate(COL_HEADERS, 1):
    cell = ws.cell(row=1, column=j, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = Border(bottom=Side(style="thin", color="999999"))
    ws.column_dimensions[get_column_letter(j)].width = max(len(str(h).split('\n')[0]) + 4, 10)
ws.row_dimensions[1].height = 30

def fmt_p(v, low=1e-4):
    if pd.isna(v): return ""
    if v < 1e-10: return "p<1e-10"
    if v < low: return f"{v:.2e}"
    return f"{v:.4f}"

def fmt_d(v, decimals=3):
    if pd.isna(v): return ""
    return f"{v:.{decimals}f}"

row = 2
for grp_name in GROUP_ORDER:
    modules = MODULE_GROUPS[grp_name]
    # Group header
    cell = ws.cell(row=row, column=1, value=grp_name.upper())
    cell.font = group_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COL_HEADERS))
    for j in range(1, len(COL_HEADERS) + 1):
        ws.cell(row=row, column=j).border = group_border
    ws.row_dimensions[row].height = 18
    row += 1

    for module in modules:
        for comp in COMP_ORDER:
            if (module, comp) not in df_lookup:
                continue
            r = df_lookup[(module, comp)]
            ci = CI_AND_POWER.get((module, comp), {})
            d_boot = ci.get('d_boot', np.nan)
            d_ci_lo = ci.get('d_ci_lo', np.nan)
            d_ci_hi = ci.get('d_ci_hi', np.nan)
            power = ci.get('power', np.nan)
            min_d = ci.get('min_detectable_d', np.nan)

            vals = [
                module.replace("_score", "").replace("_", " "),
                comp,
                fmt_d(r["H_statistic"], 2),
                fmt_p(r["KW_p"]),
                fmt_d(r["eta_squared"], 4),
                fmt_d(r["MW_U"], 1) if not pd.isna(r["MW_U"]) else "",
                fmt_p(r["MW_p"]),
                fmt_p(r["MW_p_bonf"]),
                fmt_d(d_boot, 3) if not pd.isna(d_boot) else fmt_d(r["Cohens_d"], 3),
                fmt_d(d_ci_lo, 3),
                fmt_d(d_ci_hi, 3),
                fmt_d(r["Hedges_g"], 3),
                fmt_d(r["carotid_mean"], 4),
                fmt_d(r["coronary_mean"], 4),
                fmt_d(r["femoral_mean"], 4),
                str(int(r["carotid_n"])) if not pd.isna(r["carotid_n"]) else "",
                str(int(r["coronary_n"])) if not pd.isna(r["coronary_n"]) else "",
                str(int(r["femoral_n"])) if not pd.isna(r["femoral_n"]) else "",
                f"{power:.3f}" if not pd.isna(power) else "",
                f"{min_d:.3f}" if not pd.isna(min_d) else "",
                str(r["Significance"]),
            ]
            is_sig = r["Significance"] in ("***", "**", "*")
            for j, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=j, value=val)
                cell.font = Font(name="Times New Roman", size=7,
                                 bold=is_sig and j in (4, 7, 8, 9, 21),
                                 color="CC0000" if is_sig and j == 21 else "000000")
                cell.alignment = center_align if j > 2 else left_align
                cell.border = thin_border
            ws.row_dimensions[row].height = 15
            row += 1
    row += 1  # separator

# ── Footnotes ──
row += 1
footnotes = [
    "Table S2. Effect sizes, bootstrap confidence intervals, and statistical power for module score comparisons across vascular beds.",
    "Statistical tests: Kruskal-Wallis (KW) omnibus test; Mann-Whitney U for pairwise comparisons.",
    "Effect sizes: eta-squared for KW; Cohen's d (bootstrap, n=2000) with 95% CI for pairwise MW; Hedges' g (bias-corrected d).",
    "MW p (Bonf.): Bonferroni-corrected MW p-values (3 comparisons per module). Significance: *** p<0.001, ** p<0.01, * p<0.05, ns=not significant.",
    "Achieved Power: post-hoc power for two-sample comparison given observed d and sample sizes (alpha=0.05, two-tailed).",
    "Min. Detect. d: minimum Cohen's d detectable at 80% power given sample sizes. Highlighted where observed |d| exceeds this threshold.",
    "Femoral artery n=7 donors limits power; carotid n=50, coronary n=13. Bootstrap CIs account for small-sample uncertainty.",
]
for fn in footnotes:
    cell = ws.cell(row=row, column=1, value=fn)
    cell.font = note_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COL_HEADERS))
    ws.row_dimensions[row].height = 13
    row += 1

ws.freeze_panes = "A2"
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0

# Overwrite the canonical file
out = TABLE_DIR / "TableS2_effect_sizes.xlsx"
wb.save(out)

# Also save a clean CSV of all data including new columns
all_rows = []
for grp_name in GROUP_ORDER:
    for module in MODULE_GROUPS[grp_name]:
        for comp in COMP_ORDER:
            if (module, comp) not in df_lookup:
                continue
            r = df_lookup[(module, comp)]
            ci = CI_AND_POWER.get((module, comp), {})
            row_data = {k: r[k] for k in df_old.columns}
            row_data['Cohens_d_bootstrap'] = ci.get('d_boot', np.nan)
            row_data['d_95CI_lo'] = ci.get('d_ci_lo', np.nan)
            row_data['d_95CI_hi'] = ci.get('d_ci_hi', np.nan)
            row_data['achieved_power'] = ci.get('power', np.nan)
            row_data['min_detectable_d_80pct'] = ci.get('min_detectable_d', np.nan)
            all_rows.append(row_data)
pd.DataFrame(all_rows).to_csv(TABLE_DIR / "TableS2_effect_sizes_bootstrap.csv", index=False)

print(f"Saved: {out}")
print(f"Updated: TableS2_effect_sizes_bootstrap.csv")

# ── Summary stats ──
achieved = [r['achieved_power'] for r in all_rows if not pd.isna(r.get('achieved_power', np.nan))]
detectable = sum(1 for r in all_rows
                 if not pd.isna(r.get('Cohens_d_bootstrap', np.nan))
                 and not pd.isna(r.get('min_detectable_d_80pct', np.nan))
                 and abs(r['Cohens_d_bootstrap']) > r['min_detectable_d_80pct'])
print(f"\nPower summary: median={np.median(achieved):.3f}, >0.80: {sum(1 for p in achieved if p>0.8)}/{len(achieved)}")
print(f"Comparisons exceeding min detectable d at 80% power: {detectable}/{len(all_rows)}")
print("Done.")