"""Generate Tables S1, S3, S4 for supplementary materials."""
import pandas as pd
import numpy as np
from pathlib import Path
from scipy.stats import spearmanr
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

TABLE_DIR = Path("D:/openclaw_workspace/atherosclerosis_atlas/results/supplementary")
TABLE_DIR.mkdir(exist_ok=True, parents=True)

# ── Shared styles ──
header_font = Font(name="Times New Roman", size=8, bold=True)
data_font = Font(name="Times New Roman", size=7.5)
note_font = Font(name="Times New Roman", size=7, italic=True, color="666666")
group_font = Font(name="Times New Roman", size=8, bold=True, color="1F4E79")
header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
thin_border = Border(bottom=Side(style="hair", color="CCCCCC"))
group_border = Border(bottom=Side(style="thin", color="1F4E79"))
center_align = Alignment(vertical="center", horizontal="center")
left_align = Alignment(vertical="center", horizontal="left")
wrap_align = Alignment(wrap_text=True, vertical="center", horizontal="center")

def style_header(ws, headers, row=1):
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=j, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_align
        cell.border = Border(bottom=Side(style="thin", color="999999"))
        ws.column_dimensions[get_column_letter(j)].width = max(len(str(h)) + 4, 12)
    ws.row_dimensions[row].height = 24

def style_data_row(ws, row, vals, bold_cols=None):
    bold_cols = bold_cols or set()
    for j, val in enumerate(vals, 1):
        cell = ws.cell(row=row, column=j, value=val)
        cell.font = Font(name="Times New Roman", size=7.5, bold=(j in bold_cols))
        cell.alignment = center_align if j > 1 else left_align
        cell.border = thin_border
    ws.row_dimensions[row].height = 15

def fmt_p(v, low=1e-4):
    if pd.isna(v): return ""
    if v < 1e-10: return "p<1e-10"
    if v < low: return f"{v:.2e}"
    return f"{v:.4f}"

# ═══════════════════════════════════════════
# TABLE S1: Top-50 L-R Communication Pairs
# ═══════════════════════════════════════════
print("=== Table S1: Top-50 L-R pairs ===")
lr = pd.read_csv("D:/openclaw_workspace/atherosclerosis_atlas/results/fig5/lr_communication_all_pairs.csv")
lr_top50 = lr.nlargest(50, 'comm_prob').copy()
lr_top50['rank'] = range(1, 51)
lr_top50 = lr_top50[['rank', 'sender', 'receiver', 'ligand', 'receptor', 'pair',
                       'l_frac_sender', 'r_frac_receiver', 'comm_prob']]

S1_HEADERS = ['Rank', 'Sender', 'Receiver', 'Ligand', 'Receptor', 'L-R Pair',
              'Ligand Fraction\n(Sender)', 'Receptor Fraction\n(Receiver)', 'Comm.\nProbability']

wb1 = Workbook()
ws1 = wb1.active
ws1.title = "Table S1"
style_header(ws1, S1_HEADERS)

for i, (_, r) in enumerate(lr_top50.iterrows()):
    vals = [
        r['rank'], r['sender'], r['receiver'], r['ligand'], r['receptor'],
        r['pair'], f"{r['l_frac_sender']:.4f}", f"{r['r_frac_receiver']:.4f}",
        f"{r['comm_prob']:.4f}",
    ]
    style_data_row(ws1, i + 2, vals)

# Footnotes
fn_row = len(lr_top50) + 3
footnotes = [
    "Table S1. Top 50 ligand-receptor communication pairs by geometric mean communication probability.",
    "Communication probability = sqrt(ligand_fraction_sender x receptor_fraction_receiver), computed across all sender-receiver cell-type pairs.",
    "Only L-R pairs where both ligand and receptor are expressed in >5% of cells within the respective sender/receiver populations are included.",
    "Senders: 5 macrophage subtypes; Receivers: 13 cell types (Level 1 annotation). Total significant pairs: 55 (FDR<0.05, permutation test n=1000).",
]
for fn in footnotes:
    cell = ws1.cell(row=fn_row, column=1, value=fn)
    cell.font = note_font
    ws1.merge_cells(start_row=fn_row, start_column=1, end_row=fn_row, end_column=len(S1_HEADERS))
    ws1.row_dimensions[fn_row].height = 14
    fn_row += 1

ws1.freeze_panes = "A2"
ws1.page_setup.orientation = "landscape"
out = TABLE_DIR / "TableS1_LR_top50.xlsx"
wb1.save(out)
print(f"Saved: {out}")

# ═══════════════════════════════════════════
# TABLE S3: Cross-Dataset CellChat Centrality
# ═══════════════════════════════════════════
print("=== Table S3: Cross-dataset centrality ===")
cent_cor = pd.read_csv("D:/openclaw_workspace/atherosclerosis_atlas/results/validation/gse131778_network_centrality.csv", index_col=0)
cent_car = pd.read_csv("D:/openclaw_workspace/atherosclerosis_atlas/results/validation/gse155512_network_centrality.csv", index_col=0)
cross = pd.read_csv("D:/openclaw_workspace/atherosclerosis_atlas/results/validation/gse155512_cross_dataset_centrality.csv", index_col=0)

# Common cell types
common_ct = sorted(set(cent_cor.index) & set(cent_car.index),
                   key=lambda x: cent_cor.loc[x, 'degree_centrality'] + cent_car.loc[x, 'degree_centrality'],
                   reverse=True)

wb3 = Workbook()
ws3 = wb3.active
ws3.title = "Table S3"

S3_HEADERS = [
    'Cell Type', 'Degree (Coronary)', 'Degree (Carotid)', 'Degree Ratio\n(Car/Cor)',
    'Outgoing (Cor)', 'Outgoing (Car)', 'Incoming (Cor)', 'Incoming (Car)',
    'Betweenness (Cor)', 'Betweenness (Car)',
]
style_header(ws3, S3_HEADERS)

for i, ct in enumerate(common_ct):
    cor_deg = cent_cor.loc[ct, 'degree_centrality']
    car_deg = cent_car.loc[ct, 'degree_centrality']
    ratio = car_deg / cor_deg if cor_deg > 0 else float('inf')
    vals = [
        ct,
        f"{cor_deg:.2f}", f"{car_deg:.2f}", f"{ratio:.2f}",
        f"{cent_cor.loc[ct, 'outgoing_strength']:.2f}",
        f"{cent_car.loc[ct, 'outgoing_strength']:.2f}",
        f"{cent_cor.loc[ct, 'incoming_strength']:.2f}",
        f"{cent_car.loc[ct, 'incoming_strength']:.2f}",
        f"{cent_cor.loc[ct, 'betweenness']:.4f}",
        f"{cent_car.loc[ct, 'betweenness']:.4f}",
    ]
    style_data_row(ws3, i + 2, vals)

# Spearman correlation
from scipy.stats import spearmanr
cor_degs = [cent_cor.loc[ct, 'degree_centrality'] for ct in common_ct]
car_degs = [cent_car.loc[ct, 'degree_centrality'] for ct in common_ct]
rho, pval = spearmanr(cor_degs, car_degs)

fn_row = len(common_ct) + 3
footnotes = [
    "Table S3. Cross-dataset CellChat network centrality comparison.",
    f"Common cell types: {len(common_ct)} (GSE131778 coronary artery n=11,756 cells vs GSE155512 carotid artery n=8,866 cells).",
    f"Spearman rank correlation of degree centrality (coronary vs carotid): rho = {rho:.3f}, p = {pval:.4f}.",
    "Degree centrality = weighted (in_degree + out_degree) / (n_nodes - 1) where weights = communication probability sums.",
    "Degree Ratio > 1 indicates higher centrality in carotid; < 1 indicates higher in coronary.",
]
for fn in footnotes:
    cell = ws3.cell(row=fn_row, column=1, value=fn)
    cell.font = note_font
    ws3.merge_cells(start_row=fn_row, start_column=1, end_row=fn_row, end_column=len(S3_HEADERS))
    ws3.row_dimensions[fn_row].height = 14
    fn_row += 1

ws3.freeze_panes = "A2"
ws3.page_setup.orientation = "landscape"
out = TABLE_DIR / "TableS3_cross_dataset_centrality.xlsx"
wb3.save(out)
print(f"Saved: {out}")

# ═══════════════════════════════════════════
# TABLE S4: Spearman Correlation Matrix (Metabolism vs Immune)
# ═══════════════════════════════════════════
print("=== Table S4: Spearman correlation matrix ===")
donor = pd.read_csv("D:/openclaw_workspace/atherosclerosis_atlas/results/donor_level_scores.csv")

# Define metabolic and immune module groups
METABOLIC = ['Glycolysis_score', 'OXPHOS_score', 'FAO_score', 'FAS_score',
             'Cholesterol_score', 'Hypoxia_score']
IMMUNE_INFLAM = ['TI_Inflammation_score', 'TI_PRR_score', 'TI_Metabolic_score',
                  'Acute_Inflammation_score', 'Inflammatory_Mac_score']
MACROPHAGE = ['Resident_Mac_score', 'Foamy_Mac_score', 'HMOX1_Mac_score']
TI_EPIGEN = ['TI_H3K4me3_score', 'TI_H3K27ac_score', 'TI_HDAC_SIRT_score']
PVAT = ['Healthy_PVAT_score', 'Disease_PVAT_score']
COMPOSITE = ['TI_composite']
ALL_IMMUNE = IMMUNE_INFLAM + MACROPHAGE + TI_EPIGEN + PVAT + COMPOSITE

# Compute Spearman correlations: metabolic (rows) x immune (cols)
results = []
for met in METABOLIC:
    for imm in ALL_IMMUNE:
        valid = donor[[met, imm]].dropna()
        if len(valid) < 5: continue
        rho, p = spearmanr(valid[met], valid[imm])
        # Per-bed correlations
        bed_rhos = {}
        for bed in ['carotid', 'coronary', 'femoral']:
            sub = valid.loc[donor['plaque_location'] == bed]
            if len(sub) >= 5:
                r_bed, p_bed = spearmanr(sub[met], sub[imm])
                bed_rhos[bed] = (r_bed, p_bed, len(sub))
            else:
                bed_rhos[bed] = (np.nan, np.nan, len(sub))
        results.append({
            'Metabolic_Module': met.replace('_score', ''),
            'Immune_Module': imm.replace('_score', ''),
            'Overall_rho': rho, 'Overall_p': p,
            'Overall_n': len(valid),
            'Carotid_rho': bed_rhos['carotid'][0], 'Carotid_p': bed_rhos['carotid'][1],
            'Coronary_rho': bed_rhos['coronary'][0], 'Coronary_p': bed_rhos['coronary'][1],
            'Femoral_rho': bed_rhos['femoral'][0], 'Femoral_p': bed_rhos['femoral'][1],
            'Carotid_n': bed_rhos['carotid'][2], 'Coronary_n': bed_rhos['coronary'][2],
            'Femoral_n': bed_rhos['femoral'][2],
        })

corr_df = pd.DataFrame(results)
corr_df = corr_df.sort_values('Overall_rho', ascending=False)

# Format as Excel
wb4 = Workbook()
ws4 = wb4.active
ws4.title = "Table S4"

S4_HEADERS = [
    'Metabolic\nModule', 'Immune/Inflammatory\nModule',
    'Overall rho', 'Overall p', 'N',
    'Carotid rho', 'Carotid p',
    'Coronary rho', 'Coronary p',
    'Femoral rho', 'Femoral p',
]
style_header(ws4, S4_HEADERS)

for i, (_, r) in enumerate(corr_df.iterrows()):
    vals = [
        r['Metabolic_Module'], r['Immune_Module'],
        f"{r['Overall_rho']:.3f}", fmt_p(r['Overall_p']), int(r['Overall_n']),
        f"{r['Carotid_rho']:.3f}" if not pd.isna(r['Carotid_rho']) else "n<5",
        fmt_p(r['Carotid_p']) if not pd.isna(r['Carotid_p']) else "",
        f"{r['Coronary_rho']:.3f}" if not pd.isna(r['Coronary_rho']) else "n<5",
        fmt_p(r['Coronary_p']) if not pd.isna(r['Coronary_p']) else "",
        f"{r['Femoral_rho']:.3f}" if not pd.isna(r['Femoral_rho']) else "n<5",
        fmt_p(r['Femoral_p']) if not pd.isna(r['Femoral_p']) else "",
    ]
    bold = {1, 2, 3} if r['Overall_p'] < 0.05 else {1, 2}
    style_data_row(ws4, i + 2, vals, bold_cols=bold)

# Add section breaks
row = len(corr_df) + 2
# Summary of significant correlations
sig_count = (corr_df['Overall_p'] < 0.05).sum()
total_tests = len(corr_df)
fn_row = row + 1
footnotes = [
    "Table S4. Spearman rank correlations between metabolic module scores and immune/inflammatory module scores across all donors.",
    f"Donors: carotid n=50, coronary n=13, femoral n=7. Total pairwise correlations: {total_tests}.",
    f"Significant correlations (p<0.05, uncorrected): {sig_count}/{total_tests}.",
    "Bold rows indicate p<0.05 for overall correlation. Per-bed correlations shown for exploratory analysis.",
    "Metabolic modules: Glycolysis, OXPHOS, FAO, FAS, Cholesterol, Hypoxia (MSigDB Hallmark gene sets).",
    "Immune/Inflammatory modules: trained immunity (TI) sub-scores, macrophage subtype scores, PVAT scores, epigenetic scores, and TI composite.",
]
for fn in footnotes:
    cell = ws4.cell(row=fn_row, column=1, value=fn)
    cell.font = note_font
    ws4.merge_cells(start_row=fn_row, start_column=1, end_row=fn_row, end_column=len(S4_HEADERS))
    ws4.row_dimensions[fn_row].height = 14
    fn_row += 1

ws4.freeze_panes = "A2"
ws4.page_setup.orientation = "landscape"
out = TABLE_DIR / "TableS4_spearman_metabolic_immune.xlsx"
wb4.save(out)
print(f"Saved: {out}")
print("Done.")